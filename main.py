import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl as xl

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 YaBrowser/19.4.2.702 Yowser/2.5 Safari/537.36',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8'
}
df = {}
table_pop_size = pd.read_html('http://www.statdata.ru/largest_regions_russia', header=0)
df [0] = table_pop_size[0]
df[0] = df[0].rename(columns={'Unnamed: 0': '', 'Unnamed: 2': '', 'Unnamed: 4': '', 'Unnamed: 5': ''})
# print(df[0])
writer = pd.ExcelWriter('./ТестМТС.xlsx', engine='xlsxwriter')
df[0].to_excel(writer, sheet_name="Численность населения", index=False)
####
df[1] = pd.read_excel('./osn_pok_sv.xlsx', sheet_name='1', header=5, index_col=0)
df[2] = pd.read_excel('./osn_pok_sv.xlsx', sheet_name='6', header=5, index_col=0)
df[3] = pd.read_excel('./Ikt_org(1).xlsx', sheet_name='4', header=3, index_col=0)
df[4] = pd.read_excel('./Ikt_org(1).xlsx', sheet_name='10', header=3, index_col=0)
df[1] = df[1].rename(columns={'Unnamed: 33': 'Место в РФ в 2021г.'})
# df[2] = df[2].rename(columns={'Unnamed: 0': 'Регион'})
df[1].to_excel(writer, sheet_name='Объем оказанных услуг связи')
df[2].to_excel(writer, sheet_name='Число устройств')
df[3].to_excel(writer, sheet_name='Удельный вес орг.-й, исп.-х ПК')
df[4].to_excel(writer, sheet_name='Затраты организаций')
writer.save()
#####
url = 'https://gogov.ru/articles/average-salary'
req = requests.get(url=url, headers=headers)
soup = BeautifulSoup(req.text, 'lxml')
wb = xl.load_workbook('ТестМТС.xlsx')
wb.create_sheet('Средняя зарплата')
sheet = wb['Средняя зарплата']
headers_table = soup.find('table', attrs={"id": 'm-table'}).find('thead').findAll('th')
for num, item in enumerate(headers_table):
    header = item.text
    sheet.cell(1, num + 1).value = header
    tbody = soup.find('tbody').find_all('tr')
    count = 2
for tr in tbody:
    i = 1
    for i in range(23):
        tds = tr.find_all('td')[0:][i].text
        sheet.cell(count, i+1).value = tds
    count+=1
    i+=1
wb.save('ТестМТС.xlsx')
